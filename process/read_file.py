import pandas as pd
from openpyxl import load_workbook
import xlrd
import csv
import json

'''
pandas读取excel
'''
def read_excel_by_pandas(path):
    df = pd.read_excel(path)  # pandas读取出来的格式是DataFrame
    print(df)
    # 3.读取excel的某一个sheet
    df = pd.read_excel(path, sheet_name='Sheet1')  # 读取指定sheet
    print(df)
    # 4.获取列标题
    print(df.columns)
    # 5.获取列行标题
    print(df.index)
    # 6.制定打印某一列
    print(df["姓名"])
    # 7.描述数据
    print(df.describe())
    return df

'''
openpyxl读取excel
'''
def read_excel_by_openpyxl(path):
    # 1.打开 Excel 表格并获取表格名称
    workbook = load_workbook(filename=path)  # 只支持.xlsx格式
    print(workbook.sheetnames)
    # 2.通过 sheet 名称获取表格
    sheet = workbook["Sheet1"]
    print(sheet)
    # 3.获取表格的尺寸大小(几行几列数据) 这里所说的尺寸大小，指的是 excel 表格中的数据有几行几列，针对的是不同的 sheet 而言。
    print(sheet.dimensions)
    # 4.获取表格内某个格子的数据
    # 1 sheet["A1"]方式
    cell1 = sheet["B2"]  # 按列名获取
    cell2 = sheet["C11"]
    print(cell1.value, cell2.value)
    """
    workbook.active 打开激活的表格; sheet["A1"] 获取 A1 格子的数据; cell.value 获取格子中的值;
    """
    # 4.2sheet.cell(row=, column=)方式
    cell1 = sheet.cell(row=1, column=1)
    cell2 = sheet.cell(row=11, column=3)
    print(cell1.value, cell2.value)

    # 5. 获取一系列格子
    # 获取 A1:C2 区域的值
    cell = sheet["A1:C2"]
    print(cell)
    for i in cell:
        for j in i:
            print(j.value)

'''
xlrd读取excel文件
'''
def read_excel_by_xlrd(path):
    wb = xlrd.open_workbook(path)
    # 获取并打印 sheet 数量
    print("sheet 数量:", wb.nsheets)
    # 获取并打印 sheet 名称
    print("sheet 名称:", wb.sheet_names())
    # 根据 sheet 索引获取内容
    sh1 = wb.sheet_by_index(0)
    # 也可根据 sheet 名称获取内容
    # sh = wb.sheet_by_name('成绩')
    # 获取并打印该 sheet 行数和列数
    print(u"sheet %s 共 %d 行 %d 列" % (sh1.name, sh1.nrows, sh1.ncols))
    # 获取并打印某个单元格的值
    print("第一行第二列的值为:", sh1.cell_value(0, 1))
    # 获取整行或整列的值
    rows = sh1.row_values(0)  # 获取第一行内容
    cols = sh1.col_values(1)  # 获取第二列内容
    # 打印获取的行列值
    print("第一行的值为:", rows)
    print("第二列的值为:", cols)
    # 获取单元格内容的数据类型
    print("第二行第一列的值类型为:", sh1.cell(1, 0).ctype)

'''
open函数打开csv文件
'''
def read_csv_by_open(path):
    csv_list=[]
    with open(path) as f:
        for line in f:
            list_line=line.strip('\n').split(',')  # 去除逗号和换行符号，去除后格式为list
            print(list_line)
            csv_list.append(list_line)
    print(csv_list)
    return csv_list

'''
用pandas读取csv文件
'''
def read_csv_by_pandas(path):
    # 返回的是一个DataFrame数据
    df = pd.read_csv(path)
    print(df)
    print(df.describe())
    return df
'''
csv库读取csv文件
'''
def read_csv_by_csv(path):
    csv_list=[]
    csv_reader = csv.reader(open(path))
    for line in csv_reader:
        print(line)
        csv_list.append(line)
    print(csv_list)
    return csv_list
'''
json包读取json文件
'''
def read_json_by_json(path):
    json_dict=json.load(open(path,encoding="utf-8"))
    print(json_dict)
    print(json_dict["employee"])
    return json_dict


def read_txt_all(path):
    with open(path,"r",encoding="utf-8") as file:
        data=file.read() #读取所有
    print(data)
    return data

def read_txt_to_list(path):
    res_list=[]
    with open(path, "r") as file:
        for line in file.readlines():
            line.strip("\n").split('\t')  # 去掉列表中每一个元素的换行符
            print(line)


if __name__=="__main__":
    path="../data/read_file_test/"
    file_name="number_test.txt"
    # read_txt_to_list(path+file_name)
    excel_name="test_excel.xlsx"
    #read_excel_by_xlrd(path+excel_name)
    csv_name="test_csv.csv"
    #read_csv_by_open(path+csv_name)
    #read_csv_by_pandas(path+csv_name)
    #read_csv_by_csv(path+csv_name)
    json_name="test_json.json"
    read_json_by_json(path+json_name)